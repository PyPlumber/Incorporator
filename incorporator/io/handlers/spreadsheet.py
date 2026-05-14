"""Spreadsheet format handler: Excel .xlsx via openpyxl.

Excel is the #1 business-user data format. This handler intentionally targets a
narrow, safe subset:

* **Read:** first sheet only, header row = row 1, no merged-cell expansion.
* **Write:** single sheet, headers driven by ``all_field_names`` (mirrors CSV).
* **Append mode:** not supported — Excel is a monolithic format like JSON/XML.

Anything beyond that (multi-sheet workbooks, merged cells, formatted-as-data
values, charts) is deliberately out of scope. Users with that complexity should
drop down to ``openpyxl`` directly.

openpyxl is loaded lazily inside ``parse()`` / ``write()`` so importing this
module never pulls a heavy dep at framework import time. This mirrors
``AvroHandler``'s lazy-import pattern.
"""

import logging
from pathlib import Path
from typing import Any, Dict, Iterable, List, Union

from ...exceptions import IncorporatorFormatError
from ..formats import deserialize_nested, serialize_nested
from ._base import BaseFormatHandler, _neutralise_formula_injection, _raise_if_append_unsupported, atomic_write_path

logger = logging.getLogger(__name__)


class ExcelHandler(BaseFormatHandler):
    """Parse and write .xlsx files using openpyxl.

    Lazy-imports openpyxl on first use. Raises a clear ``IncorporatorFormatError``
    pointing to ``pip install incorporator[xlsx]`` when the optional dep is
    missing.
    """

    def parse(self, source: Union[str, bytes, Path], **kwargs: Any) -> List[Dict[str, Any]]:
        """Read an ``.xlsx`` file and yield rows as dicts.

        Opens the workbook in ``read_only=True`` + ``data_only=True`` mode so
        memory stays bounded and formula cells return their evaluated values.
        Reads from the first sheet by default; pass ``sheet_name="..."`` to
        target a specific sheet. Header row is always row 1.
        """
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError:
            raise IncorporatorFormatError("openpyxl not installed. Run: pip install incorporator[xlsx]") from None

        if not isinstance(source, Path):
            raise IncorporatorFormatError("ExcelHandler requires a physical Path object.")

        try:
            # read_only=True keeps memory bounded for large workbooks; data_only=True
            # returns evaluated cell values rather than formula strings.
            wb = load_workbook(filename=source, read_only=True, data_only=True)
            try:
                sheet_name = kwargs.get("sheet_name")
                ws = wb[sheet_name] if sheet_name else wb.active
                if ws is None:
                    raise IncorporatorFormatError(f"Excel file '{source}' has no readable sheet.")

                row_iter = ws.iter_rows(values_only=True)

                try:
                    header_row = next(row_iter)
                except StopIteration:
                    return []

                # Sanitise headers: missing/blank header cells get a placeholder.
                headers: List[str] = [
                    str(h) if h is not None and str(h).strip() else f"unknown_column_{i}"
                    for i, h in enumerate(header_row)
                ]

                rows: List[Dict[str, Any]] = []
                for raw_row in row_iter:
                    # Skip fully empty rows — common with trailing whitespace in real workbooks.
                    if all(cell is None for cell in raw_row):
                        continue
                    parsed: Dict[str, Any] = {}
                    for header, cell in zip(headers, raw_row):
                        parsed[header] = deserialize_nested(cell)
                    rows.append(parsed)
                return rows
            finally:
                wb.close()
        except IncorporatorFormatError:
            raise
        except Exception as e:
            raise IncorporatorFormatError(f"Excel Read Error: {e}") from e

    def write(self, data: Iterable[Dict[str, Any]], file_path: Union[str, Path], **kwargs: Any) -> None:
        """Stream rows to an ``.xlsx`` file using openpyxl's write-only mode.

        Uses ``Workbook(write_only=True)`` so only the current row is held
        in RAM — same O(1) memory profile as ``CSVHandler``. Honours
        ``sheet_name`` (default ``"Sheet1"``) and ``all_field_names``
        (column order hint). Append mode is rejected: xlsx is a zip-of-XML
        monolith with no safe O(1) append.
        """
        # Empty guard handled centrally by _peek_iterable in handlers/__init__.py.
        # Append mode is rejected: xlsx is a zip-of-XML monolith — there is no
        # safe O(1) append. Users who need append should stream to NDJSON/CSV.
        _raise_if_append_unsupported(kwargs, "Excel/.xlsx")

        try:
            from openpyxl import Workbook
        except ImportError:
            raise IncorporatorFormatError("openpyxl not installed. Run: pip install incorporator[xlsx]") from None

        path = Path(file_path).resolve()
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        explicit_fieldnames: List[str] = kwargs.get("all_field_names") or []
        # Formula-injection mitigation defaults ON — cells starting with
        # =, @, +, -, or whitespace control chars get a single-quote prefix
        # so Excel renders them as text instead of evaluating.  Opt out via
        # ``xlsx_safe_formulas=False`` when the consumer needs raw passthrough.
        safe_formulas: bool = kwargs.get("xlsx_safe_formulas", True)

        data_iter: Iterable[Dict[str, Any]]

        if not explicit_fieldnames:
            # No schema hint (called outside export()): must materialize to discover columns.
            rows_list: List[Dict[str, Any]] = list(data)
            explicit_fieldnames = list(dict.fromkeys(k for row in rows_list for k in row))
            data_iter = iter(rows_list)
        else:
            data_iter = data

        if not explicit_fieldnames:
            return  # truly empty even after materialization

        try:
            # write_only=True streams rows directly to disk, holding only the
            # current row in memory — the equivalent of CSVHandler's generator
            # path.
            wb = Workbook(write_only=True)
            try:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(explicit_fieldnames)

                for row in data_iter:
                    # serialize_nested flattens dict/list values to JSON strings,
                    # matching CSV / SQLite behaviour. Excel has no native nested
                    # type, so this is the only safe choice.  Formula-injection
                    # mitigation prefixes any "=", "@", "+", "-" string with a
                    # single quote so Excel renders text instead of evaluating.
                    cells = [serialize_nested(row.get(k)) for k in explicit_fieldnames]
                    if safe_formulas:
                        cells = [_neutralise_formula_injection(v) for v in cells]
                    ws.append(cells)

                # Atomic save — write to a sibling tempfile then rename so
                # an interrupted save can't leave a corrupt .xlsx zip behind.
                with atomic_write_path(path) as tmp_path:
                    wb.save(str(tmp_path))
            finally:
                wb.close()
        except Exception as e:
            raise IncorporatorFormatError(f"Excel Write Error on {file_path}: {e}") from e
